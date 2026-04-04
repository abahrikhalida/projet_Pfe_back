
# from rest_framework.decorators import api_view, permission_classes, parser_classes
# from rest_framework.permissions import IsAuthenticated
# from rest_framework.parsers import MultiPartParser
# from rest_framework.response import Response
# from rest_framework import status
# from rest_framework_simplejwt.tokens import RefreshToken
# from django.utils.crypto import get_random_string
# from .models import User, Agent

# # # ==========================
# # # LOGIN API (JWT)
# # # ==========================

# @api_view(['POST'])
# def api_login(request):
#     data = request.data
#     email = data.get('email')
#     password = data.get('password')

#     try:
#         user = User.objects.get(email=email)
#     except User.DoesNotExist:
#         return Response(
#             {"status": "error", "message": "Email ou mot de passe incorrect."},
#             status=status.HTTP_401_UNAUTHORIZED
#         )

#     if not user.check_password(password):
#         return Response(
#             {"status": "error", "message": "Email ou mot de passe incorrect."},
#             status=status.HTTP_401_UNAUTHORIZED
#         )

#     refresh = RefreshToken.for_user(user)

#     return Response({
#         "status": "success",
#         "role": getattr(user, 'role', None),
#         "refresh": str(refresh),
#         "access": str(refresh.access_token),
#         "message": f"Logged in as {user.nom} {user.prenom}",
#         "nom_complet": f"{user.nom} {user.prenom}",  # <-- NOM COMPLET
#         "photo_profil": user.photo_profil.url if user.photo_profil else None
#     })

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def api_logout(request):
#     try:
#         # Invalidate refresh token
#         refresh_token = request.data.get("refresh")
#         token = RefreshToken(refresh_token)
#         token.blacklist()
#         return Response({"status": "success", "message": "Logged out successfully"})
#     except Exception as e:
#         return Response({"status": "error", "message": str(e)}, status=400)

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def api_change_password(request):
#     user = request.user
#     old_password = request.data.get("old_password")
#     new_password = request.data.get("new_password")
    
#     if not user.check_password(old_password):
#         return Response({"status": "error", "message": "Old password incorrect"}, status=400)
    
#     user.set_password(new_password)
#     user.save()
#     return Response({"status": "success", "message": "Password changed successfully"})


# from rest_framework.decorators import api_view, permission_classes
# from rest_framework.permissions import AllowAny
# from rest_framework.response import Response
# from django.utils.crypto import get_random_string
# from django.core.mail import send_mail
# from django.conf import settings
# from django.contrib.auth.tokens import PasswordResetTokenGenerator
# from .models import User
# from django.utils.http import urlsafe_base64_encode
# from django.utils.encoding import force_bytes

# @api_view(['POST'])
# @permission_classes([AllowAny])
# def api_reset_password(request):
#     email = request.data.get("email")
#     try:
#         user = User.objects.get(email=email)
#     except User.DoesNotExist:
#         return Response({"status": "error", "message": "Utilisateur non trouvé"}, status=404)
    
#     token = PasswordResetTokenGenerator().make_token(user)
#     uid = urlsafe_base64_encode(force_bytes(user.pk))
    
#     reset_url = f"http://localhost:3000/reset-confirm/{uid}/{token}"  # frontend link
#     subject = "Réinitialisation de mot de passe"
#     message = f"Bonjour {user.prenom},\n\nPour réinitialiser votre mot de passe, cliquez sur ce lien :\n{reset_url}\n\nSi vous n'avez pas demandé cette réinitialisation, ignorez cet email."
    
#     send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=False)
    
#     return Response({"status": "success", "message": "Email de réinitialisation envoyé."})

# from django.utils.http import urlsafe_base64_decode
# from django.utils.encoding import force_str
# from rest_framework import status

# @api_view(['POST'])
# @permission_classes([AllowAny])
# def api_reset_password_confirm(request):
#     uid = request.data.get("uid")
#     token = request.data.get("token")
#     new_password = request.data.get("new_password")
    
#     if not all([uid, token, new_password]):
#         return Response({"status": "error", "message": "uid, token et new_password sont requis"}, status=400)
    
#     try:
#         user_id = force_str(urlsafe_base64_decode(uid))
#         user = User.objects.get(pk=user_id)
#     except (User.DoesNotExist, ValueError):
#         return Response({"status": "error", "message": "Lien invalide"}, status=400)
    
#     if PasswordResetTokenGenerator().check_token(user, token):
#         user.set_password(new_password)
#         user.save()
#         return Response({"status": "success", "message": "Mot de passe réinitialisé avec succès"})
#     else:
#         return Response({"status": "error", "message": "Token invalide ou expiré"}, status=400)
    

# # ==========================
# # CREATE USER API
# # ==========================

# @api_view(['POST']) 
# @parser_classes([MultiPartParser])
# def api_create_user(request):
#     data = request.data
#     email = data.get('email')
#     nom = data.get('nom')
#     prenom = data.get('prenom')
#     role = data.get('role')  # <-- récupère le rôle depuis Postman
#     password = data.get('password') or get_random_string(8)
#     photo = request.FILES.get('photo_profil')

#     if not role:
#         return Response({"status": "error", "message": "Le rôle est obligatoire"}, status=400)

#     if User.objects.filter(email=email).exists():
#         return Response({"status": "error", "message": "Email déjà existant"}, status=400)

#     user = User(
#         email=email,
#         nom=nom,
#         prenom=prenom,
#         role=role,  
#         photo_profil=photo
#     )
#     user.set_password(password)
#     user.save()

#     return Response({
#         "status": "success",
#         "id": user.id,
#         "email": user.email,
#         "nom": user.nom,
#         "prenom": user.prenom,
#         "role": user.role,
#         "photo_profil": user.photo_profil.url if user.photo_profil else None,
#         "password_generated": password if 'password' not in data else None
#     })

# # ==========================
# # GET USER API
# # ==========================
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def api_get_user(request, user_id):
#     try:
#         user = User.objects.get(id=user_id)
#     except User.DoesNotExist:
#         return Response({"status": "error", "message": "User non trouvé"}, status=404)

#     return Response({
#         "id": user.id,
#         "email": user.email,
#         "nom": user.nom,
#         "prenom": user.prenom,
#         "photo_profil": user.photo_profil.url if user.photo_profil else None,
#         "is_staff": user.is_staff,
#         "is_superuser": user.is_superuser
#     })

# # -------------------------------
# # Lister tous les utilisateurs
# # -------------------------------
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def api_list_users(request):
#     users = User.objects.all()
#     data = [{
#         "id": u.id,
#         "email": u.email,
#         "nom": u.nom,
#         "prenom": u.prenom,
#         "role": u.role,
#         "photo_profil": u.photo_profil.url if u.photo_profil else None,
#         "is_staff": u.is_staff,
#         "is_superuser": u.is_superuser
#     } for u in users]

#     return Response({"status": "success", "users": data})

# # -------------------------------
# # Récupérer un utilisateur
# # -------------------------------
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def api_get_user(request, user_id):
#     try:
#         user = User.objects.get(id=user_id)
#     except User.DoesNotExist:
#         return Response({"status": "error", "message": "User non trouvé"}, status=404)

#     return Response({
#         "id": user.id,
#         "email": user.email,
#         "nom": user.nom,
#         "prenom": user.prenom,
#         "role": user.role,
#         "photo_profil": user.photo_profil.url if user.photo_profil else None,
#         "is_staff": user.is_staff,
#         "is_superuser": user.is_superuser
#     })

# # -------------------------------
# # Mettre à jour un utilisateur
# # -------------------------------
# @api_view(['PUT', 'PATCH'])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser])
# def api_update_user(request, user_id):
#     try:
#         user = User.objects.get(id=user_id)
#     except User.DoesNotExist:
#         return Response({"status": "error", "message": "User non trouvé"}, status=404)

#     data = request.data
#     user.nom = data.get('nom', user.nom)
#     user.prenom = data.get('prenom', user.prenom)
#     user.email = data.get('email', user.email)
#     user.role = data.get('role', user.role)
    
#     # Mise à jour de la photo si fournie
#     if 'photo_profil' in request.FILES:
#         user.photo_profil = request.FILES['photo_profil']

#     user.save()

#     return Response({"status": "success", "message": f"User {user.nom} mis à jour"})

# # -------------------------------
# # Supprimer un utilisateur
# # -------------------------------
# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def api_delete_user(request, user_id):
#     try:
#         user = User.objects.get(id=user_id)
#     except User.DoesNotExist:
#         return Response({"status": "error", "message": "User non trouvé"}, status=404)

#     user.delete()
#     return Response({"status": "success", "message": f"User {user.nom} supprimé"})

# # ==========================
# # CREATE AGENT API
# # ==========================


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser])
# def api_create_agent(request):
#     # Vérifie que l'utilisateur est un chef
#     if getattr(request.user, 'role', None) != 'chef':
#         return Response({"status": "error", "message": "Accès refusé"}, status=403)

#     data = request.data
#     photo = request.FILES.get('photo_profil')  # fichier image
#     password_temp = get_random_string(8)      # mot de passe temporaire

#     # Crée l'Agent
#     agent = Agent(
#         chef=request.user,
#         nom=data.get('nom'),
#         prenom=data.get('prenom'),
#         email=data.get('email'),
#         adresse=data.get('adresse'),
#         date_naissance=data.get('date_naissance'),
#         sexe=data.get('sexe'),
#         telephone=data.get('telephone'),
#         matricule=data.get('matricule'),
#         poste=data.get('poste'),
#         password_temp=password_temp,
#         is_activated=True
#     )
#     agent.save()

#     # Crée le User lié à l'agent si pas encore créé
#     if not agent.user:
#         agent.user = User.objects.create_user(
#             email=agent.email,
#             nom=agent.nom,
#             prenom=agent.prenom,
#             role='agent',
#             password=password_temp
#         )

#     # Assigne la photo au User
#     if photo:
#         agent.user.photo_profil = photo
#         agent.user.save()

#     agent.save()  # met à jour la relation user

#     # Préparer la réponse détaillée
#     response_data = {
#         "status": "success",
#         "message": f"Agent {agent.nom} créé",
#         "agent": {
#             "id": agent.id,
#             "nom": agent.nom,
#             "prenom": agent.prenom,
#             "email": agent.email,
#             "adresse": agent.adresse,
#             "date_naissance": agent.date_naissance,
#             "sexe": agent.sexe,
#             "telephone": agent.telephone,
#             "matricule": agent.matricule,
#             "poste": agent.poste,
#             "is_activated": agent.is_activated,
#             "password_temp": password_temp,
#             "user": {
#                 "id": agent.user.id,
#                 "role": agent.user.role,
#                 "photo_profil": agent.user.photo_profil.url if agent.user.photo_profil else None
#             },
#             "chef": {
#                 "id": agent.chef.id,
#                 "nom": agent.chef.nom,
#                 "prenom": agent.chef.prenom,
#                 "email": agent.chef.email,
#                 "role": agent.chef.role,
#                 "photo_profil": agent.chef.photo_profil.url if agent.chef.photo_profil else None
#             }
#         }
#     }

#     return Response(response_data)
# # ==========================
# # LIST AGENTS API
# # ==========================


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def api_list_agents(request):
#     user = request.user
#     if getattr(user, 'role', None) == 'chef':
#         agents = Agent.objects.filter(chef=user)
#     else:
#         agents = Agent.objects.all()

#     data = [{
#         "id": a.id,
#         "nom": a.nom,
#         "prenom": a.prenom,
#         "email": a.email,
#         "matricule": a.matricule,
#         "poste": a.poste,
#         "adresse": a.adresse,
#         "date_naissance": a.date_naissance.strftime('%Y-%m-%d') if a.date_naissance else None,
#         "sexe": a.sexe,
#         "telephone": a.telephone,
#         "is_activated": a.is_activated,
#         "photo_profil": a.user.photo_profil.url if a.user and a.user.photo_profil else None,
#         "chef": {  # Infos détaillées du chef
#             "id": a.chef.id if a.chef else None,
#             "nom": a.chef.nom if a.chef else None,
#             "prenom": a.chef.prenom if a.chef else None,
#             "email": a.chef.email if a.chef else None,
#             "role": a.chef.role if a.chef else None,
#             "photo_profil": a.chef.photo_profil.url if a.chef and a.chef.photo_profil else None
#         }
#     } for a in agents]

#     return Response({"status": "success", "agents": data})



# @api_view(['PUT', 'PATCH'])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser])
# def api_update_agent(request, agent_id):
#     try:
#         agent = Agent.objects.get(id=agent_id)
#     except Agent.DoesNotExist:
#         return Response({"status": "error", "message": "Agent non trouvé"}, status=404)

#     user_request = request.user
#     if getattr(user_request, 'role', None) == 'chef' and agent.chef != user_request:
#         return Response({"status": "error", "message": "Accès refusé"}, status=403)

#     data = request.data

#     # -------------------
#     # Mise à jour Agent
#     # -------------------
#     agent.nom = data.get('nom', agent.nom)
#     agent.prenom = data.get('prenom', agent.prenom)
#     agent.email = data.get('email', agent.email)
#     agent.adresse = data.get('adresse', agent.adresse)
#     agent.date_naissance = data.get('date_naissance', agent.date_naissance)
#     agent.sexe = data.get('sexe', agent.sexe)
#     agent.telephone = data.get('telephone', agent.telephone)
#     agent.matricule = data.get('matricule', agent.matricule)
#     agent.poste = data.get('poste', agent.poste)  # <-- ajout du poste
#     agent.save()

#     # -------------------
#     # Mise à jour User lié (role + photo)
#     # -------------------
#     photo = request.FILES.get('photo_profil')
#     if agent.user:
#         agent.user.nom = agent.nom
#         agent.user.prenom = agent.prenom
#         agent.user.email = agent.email
#         if photo:
#             agent.user.photo_profil = photo
#         agent.user.save()

#     # -------------------
#     # Construire la réponse détaillée
#     # -------------------
#     response_data = {
#         "id": agent.id,
#         "chef": {
#             "id": agent.chef.id,
#             "nom": agent.chef.nom,
#             "prenom": agent.chef.prenom,
#             "email": agent.chef.email,
#             "role": agent.chef.role,
#             "photo_profil": agent.chef.photo_profil.url if agent.chef.photo_profil else None
#         },
#         "nom": agent.nom,
#         "prenom": agent.prenom,
#         "email": agent.email,
#         "adresse": agent.adresse,
#         "date_naissance": agent.date_naissance,
#         "sexe": agent.sexe,
#         "telephone": agent.telephone,
#         "matricule": agent.matricule,
#         "poste": agent.poste,  # <-- ajout du poste dans la réponse
#         "password_temp": agent.password_temp,
#         "activation_code": agent.activation_code,
#         "is_activated": agent.is_activated,
#         "role": agent.user.role if agent.user else None,
#         "photo_profil": agent.user.photo_profil.url if agent.user and agent.user.photo_profil else None
#     }

#     return Response({
#         "status": "success",
#         "message": f"Agent {agent.nom} mis à jour",
#         "agent": response_data
#     })
# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def api_delete_agent(request, agent_id):
#     try:
#         agent = Agent.objects.get(id=agent_id)
#     except Agent.DoesNotExist:
#         return Response({"status": "error", "message": "Agent non trouvé"}, status=404)

#     user = request.user
#     if getattr(user, 'role', None) == 'chef' and agent.chef != user:
#         return Response({"status": "error", "message": "Accès refusé"}, status=403)

#     agent.delete()
#     return Response({"status": "success", "message": f"Agent {agent.nom} supprimé"})



# from rest_framework.decorators import api_view, permission_classes, parser_classes
# from rest_framework.permissions import IsAuthenticated
# from rest_framework.parsers import MultiPartParser
# from rest_framework.response import Response
# from datetime import datetime, date, timedelta
# import openpyxl
# from django.utils.crypto import get_random_string
# from api.models import Agent  
# from django.core.mail import send_mail
# from django.conf import settings


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser])
# def api_create_agents_excel(request):
#     """
#     Crée plusieurs agents à partir d'un fichier Excel (.xlsx)
#     Format attendu : nom | prenom | email | adresse | date_naissance | sexe | telephone | matricule | poste
#     """
#     if getattr(request.user, 'role', None) != 'chef':
#         return Response({"status": "error", "message": "Accès refusé"}, status=403)

#     file_obj = request.FILES.get('file')
#     if not file_obj:
#         return Response({"status": "error", "message": "Fichier Excel requis"}, status=400)

#     try:
#         wb = openpyxl.load_workbook(file_obj, data_only=True)
#         sheet = wb.active
#     except Exception as e:
#         return Response({"status": "error", "message": f"Impossible de lire le fichier : {str(e)}"}, status=400)

#     created_agents = []
#     skipped_agents = []

#     for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#         if not any(row):
#             continue
#         if len(row) < 9:
#             skipped_agents.append({"row": row_idx, "reason": "Colonnes manquantes"})
#             continue

#         nom, prenom, email, adresse, date_naissance, sexe, telephone, matricule, poste = row
#         email = str(email).strip() if email else None
#         if not email:
#             skipped_agents.append({"row": row_idx, "reason": "Email manquant"})
#             continue
#         if Agent.objects.filter(email=email).exists():
#             skipped_agents.append({"email": email, "reason": "Email déjà existant"})
#             continue

#         # Gestion date
#         dn = None
#         if isinstance(date_naissance, (datetime, date)):
#             dn = date_naissance if isinstance(date_naissance, date) else date_naissance.date()
#         elif isinstance(date_naissance, str):
#             for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
#                 try:
#                     dn = datetime.strptime(date_naissance.strip(), fmt).date()
#                     break
#                 except ValueError:
#                     continue
#         elif isinstance(date_naissance, (int, float)):
#             dn = (date(1899, 12, 30) + timedelta(days=int(date_naissance)))

#         if dn is None:
#             skipped_agents.append({"email": email, "reason": f"Format date invalide: {date_naissance}"} )
#             continue

#         try:
#             # Création agent ; le save() gère User + email
#             agent = Agent(
#                 chef=request.user,
#                 nom=str(nom).strip() if nom else "",
#                 prenom=str(prenom).strip() if prenom else "",
#                 email=email,
#                 adresse=str(adresse).strip() if adresse else "",
#                 date_naissance=dn,
#                 sexe=str(sexe).strip().upper() if sexe else "",
#                 telephone=str(telephone).strip() if telephone else "",
#                 matricule=str(matricule).strip() if matricule else "",
#                 poste=str(poste).strip() if poste else "",
#             )
#             agent.save()  # <-- save() crée User + envoie email

#             created_agents.append({
#                 "id": agent.id,
#                 "nom": agent.nom,
#                 "prenom": agent.prenom,
#                 "email": agent.email,
#                 "adresse": agent.adresse,
#                 "date_naissance": agent.date_naissance.strftime('%Y-%m-%d') if agent.date_naissance else None,
#                 "sexe": agent.sexe,
#                 "telephone": agent.telephone,
#                 "matricule": agent.matricule,
#                 "poste": agent.poste,
#                 "password_temp": agent.password_temp,
#                 "activation_code": agent.activation_code,
#                 "is_activated": agent.is_activated,
#                 "user": {
#                     "id": agent.user.id if agent.user else None,
#                     "role": agent.user.role if agent.user else None,
#                     "photo_profil": agent.user.photo_profil.url if agent.user and agent.user.photo_profil else None
#                 },
#                 "chef": {
#                     "id": agent.chef.id if agent.chef else None,
#                     "nom": agent.chef.nom if agent.chef else None,
#                     "prenom": agent.chef.prenom if agent.chef else None,
#                     "email": agent.chef.email if agent.chef else None,
#                     "role": agent.chef.role if agent.chef else None,
#                     "photo_profil": agent.chef.photo_profil.url if agent.chef and agent.chef.photo_profil else None
#                 }
#             })

#         except Exception as e:
#             skipped_agents.append({"email": email, "reason": str(e)})

#     # Résumé final
#     return Response({
#         "status": "success",
#         "total_created": len(created_agents),
#         "created": created_agents,
#         "skipped": skipped_agents
#     })
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from django.utils.crypto import get_random_string
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from datetime import datetime, date, timedelta
import openpyxl
from .models import User, Agent


# Helper pour nom complet
def nom_complet(obj):
    return f"{obj.prenom} {obj.nom}"


# ==========================
# LOGIN API (JWT)
# ==========================

@api_view(['POST'])
def api_login(request):
    data = request.data
    email = data.get('email')
    password = data.get('password')

    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response(
            {"status": "error", "message": "Email ou mot de passe incorrect."},
            status=status.HTTP_401_UNAUTHORIZED
        )

    if not user.check_password(password):
        return Response(
            {"status": "error", "message": "Email ou mot de passe incorrect."},
            status=status.HTTP_401_UNAUTHORIZED
        )

    refresh = RefreshToken.for_user(user)

    return Response({
        "status": "success",
        "role": getattr(user, 'role', None),
        "refresh": str(refresh),
        "access": str(refresh.access_token),
        "message": f"Logged in as {nom_complet(user)}",
        "nom_complet": nom_complet(user),
        "photo_profil": user.photo_profil.url if user.photo_profil else None
    })


# ==========================
# LOGOUT
# ==========================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_logout(request):
    try:
        refresh_token = request.data.get("refresh")
        token = RefreshToken(refresh_token)
        token.blacklist()
        return Response({"status": "success", "message": "Logged out successfully"})
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=400)


# ==========================
# CHANGE PASSWORD
# ==========================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_change_password(request):
    user = request.user
    old_password = request.data.get("old_password")
    new_password = request.data.get("new_password")

    if not user.check_password(old_password):
        return Response({"status": "error", "message": "Old password incorrect"}, status=400)

    user.set_password(new_password)
    user.save()
    return Response({"status": "success", "message": "Password changed successfully"})


# ==========================
# RESET PASSWORD
# ==========================

@api_view(['POST'])
@permission_classes([AllowAny])
def api_reset_password(request):
    email = request.data.get("email")
    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response({"status": "error", "message": "Utilisateur non trouvé"}, status=404)

    token = PasswordResetTokenGenerator().make_token(user)
    uid = urlsafe_base64_encode(force_bytes(user.pk))

    reset_url = f"http://localhost:3000/reset-confirm/{uid}/{token}"
    subject = "Réinitialisation de mot de passe"
    message = f"Bonjour {user.prenom},\n\nPour réinitialiser votre mot de passe, cliquez sur ce lien :\n{reset_url}\n\nSi vous n'avez pas demandé cette réinitialisation, ignorez cet email."

    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=False)

    return Response({"status": "success", "message": "Email de réinitialisation envoyé."})


@api_view(['POST'])
@permission_classes([AllowAny])
def api_reset_password_confirm(request):
    uid = request.data.get("uid")
    token = request.data.get("token")
    new_password = request.data.get("new_password")

    if not all([uid, token, new_password]):
        return Response({"status": "error", "message": "uid, token et new_password sont requis"}, status=400)

    try:
        user_id = force_str(urlsafe_base64_decode(uid))
        user = User.objects.get(pk=user_id)
    except (User.DoesNotExist, ValueError):
        return Response({"status": "error", "message": "Lien invalide"}, status=400)

    if PasswordResetTokenGenerator().check_token(user, token):
        user.set_password(new_password)
        user.save()
        return Response({"status": "success", "message": "Mot de passe réinitialisé avec succès"})
    else:
        return Response({"status": "error", "message": "Token invalide ou expiré"}, status=400)


# ==========================
# CREATE USER
# ==========================

@api_view(['POST'])
@parser_classes([MultiPartParser])
def api_create_user(request):
    data = request.data
    email = data.get('email')
    nom = data.get('nom')
    prenom = data.get('prenom')
    role = data.get('role')
    password = data.get('password') or get_random_string(8)
    photo = request.FILES.get('photo_profil')

    if not role:
        return Response({"status": "error", "message": "Le rôle est obligatoire"}, status=400)

    if User.objects.filter(email=email).exists():
        return Response({"status": "error", "message": "Email déjà existant"}, status=400)

    user = User(email=email, nom=nom, prenom=prenom, role=role, photo_profil=photo)
    user.set_password(password)
    user.save()

    return Response({
        "status": "success",
        "id": user.id,
        "email": user.email,
        "nom_complet": nom_complet(user),
        "role": user.role,
        "photo_profil": user.photo_profil.url if user.photo_profil else None,
        "password_generated": password if 'password' not in data else None
    })


# ==========================
# LIST USERS
# ==========================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_list_users(request):
    users = User.objects.all()
    data = [{
        "id": u.id,
        "email": u.email,
        "nom_complet": nom_complet(u),
        "role": u.role,
        "photo_profil": u.photo_profil.url if u.photo_profil else None,
        "is_staff": u.is_staff,
        "is_superuser": u.is_superuser
    } for u in users]

    return Response({"status": "success", "users": data})


# ==========================
# GET USER
# ==========================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_get_user(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"status": "error", "message": "User non trouvé"}, status=404)

    return Response({
        "id": user.id,
        "email": user.email,
        "nom_complet": nom_complet(user),
        "role": user.role,
        "photo_profil": user.photo_profil.url if user.photo_profil else None,
        "is_staff": user.is_staff,
        "is_superuser": user.is_superuser
    })


# ==========================
# UPDATE USER
# ==========================

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def api_update_user(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"status": "error", "message": "User non trouvé"}, status=404)

    data = request.data
    user.nom = data.get('nom', user.nom)
    user.prenom = data.get('prenom', user.prenom)
    user.email = data.get('email', user.email)
    user.role = data.get('role', user.role)

    if 'photo_profil' in request.FILES:
        user.photo_profil = request.FILES['photo_profil']

    user.save()

    return Response({
        "status": "success",
        "message": f"User {nom_complet(user)} mis à jour",
        "user": {
            "id": user.id,
            "email": user.email,
            "nom_complet": nom_complet(user),
            "role": user.role,
            "photo_profil": user.photo_profil.url if user.photo_profil else None,
        }
    })


# ==========================
# DELETE USER
# ==========================

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_user(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"status": "error", "message": "User non trouvé"}, status=404)

    user.delete()
    return Response({"status": "success", "message": f"User {nom_complet(user)} supprimé"})


# ==========================
# CREATE AGENT
# ==========================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def api_create_agent(request):
    if getattr(request.user, 'role', None) != 'chef':
        return Response({"status": "error", "message": "Accès refusé"}, status=403)

    data = request.data
    photo = request.FILES.get('photo_profil')
    password_temp = get_random_string(8)

    agent = Agent(
        chef=request.user,
        nom=data.get('nom'),
        prenom=data.get('prenom'),
        email=data.get('email'),
        adresse=data.get('adresse'),
        date_naissance=data.get('date_naissance'),
        sexe=data.get('sexe'),
        telephone=data.get('telephone'),
        matricule=data.get('matricule'),
        poste=data.get('poste'),
        password_temp=password_temp,
        is_activated=True
    )
    agent.save()

    if not agent.user:
        agent.user = User.objects.create_user(
            email=agent.email,
            nom=agent.nom,
            prenom=agent.prenom,
            role='agent',
            password=password_temp
        )

    if photo:
        agent.user.photo_profil = photo
        agent.user.save()

    agent.save()

    return Response({
        "status": "success",
        "message": f"Agent {nom_complet(agent)} créé",
        "agent": {
            "id": agent.id,
            "nom_complet": nom_complet(agent),
            "email": agent.email,
            "adresse": agent.adresse,
            "date_naissance": agent.date_naissance,
            "sexe": agent.sexe,
            "telephone": agent.telephone,
            "matricule": agent.matricule,
            "poste": agent.poste,
            "is_activated": agent.is_activated,
            "password_temp": password_temp,
            "user": {
                "id": agent.user.id,
                "role": agent.user.role,
                "photo_profil": agent.user.photo_profil.url if agent.user.photo_profil else None
            },
            "chef": {
                "id": agent.chef.id,
                "nom_complet": nom_complet(agent.chef),
                "email": agent.chef.email,
                "role": agent.chef.role,
                "photo_profil": agent.chef.photo_profil.url if agent.chef.photo_profil else None
            }
        }
    })


# ==========================
# LIST AGENTS
# ==========================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_list_agents(request):
    user = request.user
    agents = Agent.objects.filter(chef=user) if getattr(user, 'role', None) == 'chef' else Agent.objects.all()

    data = [{
        "id": a.id,
        "nom_complet": nom_complet(a),
        "email": a.email,
        "matricule": a.matricule,
        "poste": a.poste,
        "adresse": a.adresse,
        "date_naissance": a.date_naissance.strftime('%Y-%m-%d') if a.date_naissance else None,
        "sexe": a.sexe,
        "telephone": a.telephone,
        "is_activated": a.is_activated,
        "photo_profil": a.user.photo_profil.url if a.user and a.user.photo_profil else None,
        "chef": {
            "id": a.chef.id if a.chef else None,
            "nom_complet": nom_complet(a.chef) if a.chef else None,
            "email": a.chef.email if a.chef else None,
            "role": a.chef.role if a.chef else None,
            "photo_profil": a.chef.photo_profil.url if a.chef and a.chef.photo_profil else None
        }
    } for a in agents]

    return Response({"status": "success", "agents": data})


# ==========================
# UPDATE AGENT
# ==========================


# le chef modifier profil de agent 

# @api_view(['PUT', 'PATCH'])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser])
# def api_update_agent(request, agent_id):
#     try:
#         agent = Agent.objects.get(id=agent_id)
#     except Agent.DoesNotExist:
#         return Response({"status": "error", "message": "Agent non trouvé"}, status=404)

#     user_request = request.user
#     if getattr(user_request, 'role', None) == 'chef' and agent.chef != user_request:
#         return Response({"status": "error", "message": "Accès refusé"}, status=403)

#     data = request.data
#     agent.nom = data.get('nom', agent.nom)
#     agent.prenom = data.get('prenom', agent.prenom)
#     agent.email = data.get('email', agent.email)
#     agent.adresse = data.get('adresse', agent.adresse)
#     agent.date_naissance = data.get('date_naissance', agent.date_naissance)
#     agent.sexe = data.get('sexe', agent.sexe)
#     agent.telephone = data.get('telephone', agent.telephone)
#     agent.matricule = data.get('matricule', agent.matricule)
#     agent.poste = data.get('poste', agent.poste)
#     agent.save()

#     photo = request.FILES.get('photo_profil')
#     if agent.user:
#         agent.user.nom = agent.nom
#         agent.user.prenom = agent.prenom
#         agent.user.email = agent.email
#         if photo:
#             agent.user.photo_profil = photo
#         agent.user.save()

#     return Response({
#         "status": "success",
#         "message": f"Agent {nom_complet(agent)} mis à jour",
#         "agent": {
#             "id": agent.id,
#             "nom_complet": nom_complet(agent),
#             "email": agent.email,
#             "adresse": agent.adresse,
#             "date_naissance": agent.date_naissance,
#             "sexe": agent.sexe,
#             "telephone": agent.telephone,
#             "matricule": agent.matricule,
#             "poste": agent.poste,
#             "password_temp": agent.password_temp,
#             "activation_code": agent.activation_code,
#             "is_activated": agent.is_activated,
#             "role": agent.user.role if agent.user else None,
#             "photo_profil": agent.user.photo_profil.url if agent.user and agent.user.photo_profil else None,
#             "chef": {
#                 "id": agent.chef.id,
#                 "nom_complet": nom_complet(agent.chef),
#                 "email": agent.chef.email,
#                 "role": agent.chef.role,
#                 "photo_profil": agent.chef.photo_profil.url if agent.chef.photo_profil else None
#             }
#         }
#     })


# le chef modifier profil de agent  +agent modifier son profil
@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def api_update_agent(request, agent_id):
    try:
        agent = Agent.objects.get(id=agent_id)
    except Agent.DoesNotExist:
        return Response({"status": "error", "message": "Agent non trouvé"}, status=404)

    user_request = request.user

    # Chef ne peut modifier que ses propres agents
    if getattr(user_request, 'role', None) == 'chef' and agent.chef != user_request:
        return Response({"status": "error", "message": "Accès refusé"}, status=403)

    # Agent ne peut modifier que son propre profil
    if getattr(user_request, 'role', None) == 'agent' and agent.user != user_request:
        return Response({"status": "error", "message": "Accès refusé"}, status=403)

    data = request.data

    # Agent ne peut pas modifier ces champs sensibles
    if getattr(user_request, 'role', None) == 'agent':
        agent.nom = data.get('nom', agent.nom)
        agent.prenom = data.get('prenom', agent.prenom)
        agent.adresse = data.get('adresse', agent.adresse)
        agent.telephone = data.get('telephone', agent.telephone)
    else:
        # Chef/admin peut tout modifier
        agent.nom = data.get('nom', agent.nom)
        agent.prenom = data.get('prenom', agent.prenom)
        agent.email = data.get('email', agent.email)
        agent.adresse = data.get('adresse', agent.adresse)
        agent.date_naissance = data.get('date_naissance', agent.date_naissance)
        agent.sexe = data.get('sexe', agent.sexe)
        agent.telephone = data.get('telephone', agent.telephone)
        agent.matricule = data.get('matricule', agent.matricule)
        agent.poste = data.get('poste', agent.poste)

    agent.save()

    photo = request.FILES.get('photo_profil')
    if agent.user:
        agent.user.nom = agent.nom
        agent.user.prenom = agent.prenom
        agent.user.email = agent.email
        if photo:
            agent.user.photo_profil = photo
        agent.user.save()

    return Response({
        "status": "success",
        "message": f"Agent {nom_complet(agent)} mis à jour",
        "agent": {
            "id": agent.id,
            "nom_complet": nom_complet(agent),
            "email": agent.email,
            "adresse": agent.adresse,
            "date_naissance": agent.date_naissance,
            "sexe": agent.sexe,
            "telephone": agent.telephone,
            "matricule": agent.matricule,
            "poste": agent.poste,
            "password_temp": agent.password_temp,
            "activation_code": agent.activation_code,
            "is_activated": agent.is_activated,
            "role": agent.user.role if agent.user else None,
            "photo_profil": agent.user.photo_profil.url if agent.user and agent.user.photo_profil else None,
            "chef": {
                "id": agent.chef.id,
                "nom_complet": nom_complet(agent.chef),
                "email": agent.chef.email,
                "role": agent.chef.role,
                "photo_profil": agent.chef.photo_profil.url if agent.chef.photo_profil else None
            }
        }
    })

# ==========================
# DELETE AGENT
# ==========================

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_agent(request, agent_id):
    try:
        agent = Agent.objects.get(id=agent_id)
    except Agent.DoesNotExist:
        return Response({"status": "error", "message": "Agent non trouvé"}, status=404)

    user = request.user
    if getattr(user, 'role', None) == 'chef' and agent.chef != user:
        return Response({"status": "error", "message": "Accès refusé"}, status=403)

    agent.delete()
    return Response({"status": "success", "message": f"Agent {nom_complet(agent)} supprimé"})


# ==========================
# CREATE AGENTS FROM EXCEL
# ==========================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def api_create_agents_excel(request):
    if getattr(request.user, 'role', None) != 'chef':
        return Response({"status": "error", "message": "Accès refusé"}, status=403)

    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({"status": "error", "message": "Fichier Excel requis"}, status=400)

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
    except Exception as e:
        return Response({"status": "error", "message": f"Impossible de lire le fichier : {str(e)}"}, status=400)

    created_agents = []
    skipped_agents = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        if len(row) < 9:
            skipped_agents.append({"row": row_idx, "reason": "Colonnes manquantes"})
            continue

        nom, prenom, email, adresse, date_naissance, sexe, telephone, matricule, poste = row
        email = str(email).strip() if email else None
        if not email:
            skipped_agents.append({"row": row_idx, "reason": "Email manquant"})
            continue
        if Agent.objects.filter(email=email).exists():
            skipped_agents.append({"email": email, "reason": "Email déjà existant"})
            continue

        dn = None
        if isinstance(date_naissance, (datetime, date)):
            dn = date_naissance if isinstance(date_naissance, date) else date_naissance.date()
        elif isinstance(date_naissance, str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                try:
                    dn = datetime.strptime(date_naissance.strip(), fmt).date()
                    break
                except ValueError:
                    continue
        elif isinstance(date_naissance, (int, float)):
            dn = (date(1899, 12, 30) + timedelta(days=int(date_naissance)))

        if dn is None:
            skipped_agents.append({"email": email, "reason": f"Format date invalide: {date_naissance}"})
            continue

        try:
            agent = Agent(
                chef=request.user,
                nom=str(nom).strip() if nom else "",
                prenom=str(prenom).strip() if prenom else "",
                email=email,
                adresse=str(adresse).strip() if adresse else "",
                date_naissance=dn,
                sexe=str(sexe).strip().upper() if sexe else "",
                telephone=str(telephone).strip() if telephone else "",
                matricule=str(matricule).strip() if matricule else "",
                poste=str(poste).strip() if poste else "",
            )
            agent.save()

            created_agents.append({
                "id": agent.id,
                "nom_complet": nom_complet(agent),
                "email": agent.email,
                "adresse": agent.adresse,
                "date_naissance": agent.date_naissance.strftime('%Y-%m-%d') if agent.date_naissance else None,
                "sexe": agent.sexe,
                "telephone": agent.telephone,
                "matricule": agent.matricule,
                "poste": agent.poste,
                "password_temp": agent.password_temp,
                "activation_code": agent.activation_code,
                "is_activated": agent.is_activated,
                "user": {
                    "id": agent.user.id if agent.user else None,
                    "role": agent.user.role if agent.user else None,
                    "photo_profil": agent.user.photo_profil.url if agent.user and agent.user.photo_profil else None
                },
                "chef": {
                    "id": agent.chef.id if agent.chef else None,
                    "nom_complet": nom_complet(agent.chef) if agent.chef else None,
                    "email": agent.chef.email if agent.chef else None,
                    "role": agent.chef.role if agent.chef else None,
                    "photo_profil": agent.chef.photo_profil.url if agent.chef and agent.chef.photo_profil else None
                }
            })

        except Exception as e:
            skipped_agents.append({"email": email, "reason": str(e)})

    return Response({
        "status": "success",
        "total_created": len(created_agents),
        "created": created_agents,
        "skipped": skipped_agents
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_me(request):
    user = request.user
    return Response({
        'id': user.id,
        'email': user.email,
        'role': getattr(user, 'role', None),
        'nom_complet': f"{user.prenom} {user.nom}",
        'is_active': user.is_active,
    })